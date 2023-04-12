#!/usr/bin/env python
# _*_ coding:utf-8 _*_
# __author__ = 'superfeng'
import openpyxl

def get_test_data_from_excel(file,sheet):
    """
    从excel读取测试数据
    :param file:
    :param sheet:
    :return:
    """
    #打开工作簿，获取工作表
    wb = openpyxl.load_workbook(filename=file)
    ws =wb[sheet]

    #获取最大行，最大列
    column =ws.max_column
    row = ws.max_row

    #读取数据
    data =[]

    #读取所有的标题key
    keys =[]
    for i in (1,column+1):
        keys.append(ws.cell(1,i).value)

    #读取所有的值
    for i in (2, row+1):
        temp ={}
        for j in (1,column+1):
            # key = keys[j-1]
            # value = ws.cell(i,j).value
            temp[keys[j-1]] = ws.cell(i,j).value
        data.append(temp)

    return data




